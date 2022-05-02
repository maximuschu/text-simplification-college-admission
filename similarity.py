from nltk.translate.bleu_score import sentence_bleu
import nltk
import openpyxl
import pandas as pd
from sentence_transformers import SentenceTransformer, util
import random
file_numbers = [4, 10, 18, 20, 22, 23, 24, 25, 26, 30, 31, 32, 34, 37, 38, 45, 46, 48, 49, 50, 51, 52, 53, 56, 58, 59, 61, 65, 71, 73, 76, 77, 85, 88, 91, 94, 97, 98, 100, 101, 102, 103, 105, 106, 107, 108, 109, 110, 114, 115, 119, 121, 123, 127, 129, 130, 140, 142, 143, 149, 156, 157, 163, 164, 166, 170, 171, 172, 174, 175, 180, 184, 191, 192, 199, 200, 211, 213, 215, 216, 217, 218, 221, 223, 224, 227, 229, 232, 235, 237, 241, 243, 246, 247, 250, 253, 258, 259, 272, 273, 278, 279, 286,
                291, 301, 308, 309, 310, 311, 317, 325, 330]
filter_by_length = False
filter_length = 5
filter_out_info = True


# Calculates BLEU scores for all files directly from list of files, stores in a single spreadsheet.
def bleu_from_files():
    nltk.download('punkt')
    column = 1
    path = "BLEU Results.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    total_result = open(
        f"BLEU/Generated and Simplified/Total.txt", "w", encoding='utf-8')
    total_average = 0
    last_column = len(file_numbers) + 1
    last_row = 2
    cell_obj = sheet_obj.cell(row=1, column=last_column)
    cell_obj.value = "File Totals"
    for file_number in file_numbers:
        file = open(
            f"Output/Sentences/{file_number}.txt", "r", encoding='utf-8')
        result = open(
            f"BLEU/Generated and Simplified/{file_number}.txt", "w", encoding='utf-8')
        current_aligning = None
        current_original = None
        current_simple = None
        current_generated = None
        average = 0
        total = 0
        cell_obj = sheet_obj.cell(row=1, column=column)
        cell_obj.value = "File " + str(file_numbers[column-1])
        row = 2
        for index, line in enumerate(file):
            if(index % 5 == 0):
                current_aligning = line
            if(index % 5 == 1):
                current_original = line
                if(filter_out_info == False):
                    current_original = current_original.replace('<INFO>', '')
            if(index % 5 == 2):
                current_simple = line
                if(filter_out_info == False):
                    current_simple = current_simple.replace('<INFO>', '')
            if(index % 5 == 3):
                current_generated = line
                if(filter_out_info == False):
                    current_generated = current_generated.replace('<INFO>', '')
            if(index % 5 == 4):
                ref = []
                test = current_simple.split()
                if("." in current_generated):
                    sentences = nltk.tokenize.sent_tokenize(current_generated)
                    for current in sentences:
                        ref.append(current.split())
                else:
                    ref.append(current_generated.split())
                current_score = sentence_bleu(ref, test)
                if(filter_out_info == True and "<INFO>" in current_simple):
                    continue
                if(filter_by_length == True and len(test) <= filter_length):
                    continue
                result.write(current_aligning)
                cell_obj = sheet_obj.cell(row=row, column=column)
                cell_obj.value = str(current_aligning)
                row = row + 1
                result.write(str(current_score) + "\n")
                cell_obj = sheet_obj.cell(row=row, column=column)
                cell_obj.value = str(current_score)
                row = row + 1
                total = total + 1
                average = average + current_score

        average = average / total
        result.write("Total: " + str(average))
        cell_obj = sheet_obj.cell(row=row, column=column)
        cell_obj.value = "Total: " + str(average)
        column = column + 1
        total_average = total_average + average
        total_result.write(str(file_number) + ": " + str(average) + "\n")
        cell_obj = sheet_obj.cell(row=last_row, column=last_column)
        cell_obj.value = str(file_number) + ": " + str(average)
        last_row = last_row + 1

    total_average = total_average / len(file_numbers)
    total_result.write("Total: " + str(total_average))
    cell_obj = sheet_obj.cell(row=last_row, column=last_column)
    cell_obj.value = "Total: " + str(total_average)
    wb_obj.save("BLEU Results.xlsx")


# Calculates both BLEU and Sentence-BERT scores for all files from single computed csv file from create_csv(), stores in a text file.
def bleu_and_bert():
    model = SentenceTransformer('paraphrase-MiniLM-L6-v2')
    result = open(
        f"BLEU/Access/Access BLEU Total.txt", "w", encoding='utf-8')
    bert_result = open(
        f"Access BERT Total.txt", "w", encoding='utf-8')
    average = 0
    bert_average = 0

    df = pd.read_csv("Models/T5/access_final.csv")
    df = df.drop('Unnamed: 0', 1)

    generated = df['Generated Text'].tolist()
    actual = df['Actual Text'].tolist()
    for i in range(len(generated)):
        ref = []
        test = actual[i].split()
        ref.append(generated[i].split())
        current_score = sentence_bleu(ref, test)
        emb1 = model.encode(generated[i])
        emb2 = model.encode(actual[i])

        bert_score = util.cos_sim(emb1, emb2).item()
        result.write("Generated: " + generated[i] + "\n")
        result.write("Actual: " + actual[i] + "\n")
        result.write(str(current_score) + "\n\n")
        average = average + current_score
        bert_result.write("Generated: " + generated[i] + "\n")
        bert_result.write("Actual: " + actual[i] + "\n")
        bert_result.write(str(bert_score) + "\n\n")
        bert_average = bert_average + bert_score
    average = average / len(generated)
    bert_average = bert_average / len(generated)
    result.write("Total: " + str(average))
    bert_result.write("Total: " + str(bert_average))


# Gathers the sentences from all files to be used for testing, outputs into a single csv file.
def create_csv():
    original = []
    actual = []
    generated = []
    random.seed(18)
    random.shuffle(file_numbers)

    file_numbers = file_numbers[int(
        len(file_numbers)*0.5):int(len(file_numbers)*0.8)]
    for file_number in file_numbers:
        file = open(
            f"Output/Final/{file_number}.txt", "r", encoding='utf-8')
        for index, line in enumerate(file):
            line = line.strip().replace("\n", "")
            if(index % 5 == 0):
                pass
            if(index % 5 == 1):
                if(filter_out_info == True and "<INFO>" in line):
                    continue
                original.append(line)
            if(index % 5 == 2):
                if(filter_out_info == True and "<INFO>" in line):
                    continue
                actual.append(line)
            if(index % 5 == 3):
                if(filter_out_info == True and "<INFO>" in line):
                    continue
                generated.append(line)
            if(index % 5 == 4):
                pass
    for i in range(len(original) - 1, -1, -1):
        if(len(original[i]) == 0):
            del original[i]
            del actual[i]
            del generated[i]

    for i in range(len(actual) - 1, -1, -1):
        if(len(actual[i]) == 0):
            del original[i]
            del actual[i]
            del generated[i]

    final_df = pd.DataFrame(
        {'Original Text': original, 'Generated Text': generated, 'Actual Text': actual})
    final_df.to_csv('access.csv')
    print(final_df.values)


def main():
    # Uncomment method for whenever function is needed.
    # bleu_from_files()
    bleu_and_bert()
    # create_csv()


if __name__ == '__main__':
    main()
